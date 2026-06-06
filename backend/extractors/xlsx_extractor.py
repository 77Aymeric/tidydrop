from __future__ import annotations

from pathlib import Path

from backend.config import MAX_PREVIEW_CHARS


def extract_xlsx(path: Path) -> tuple[str, str, str, str | None]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        lines = [f"Sheets: {', '.join(workbook.sheetnames)}"]
        for sheet in workbook.worksheets[:5]:
            lines.append(f"[{sheet.title}]")
            for row in sheet.iter_rows(min_row=1, max_row=8, max_col=8, values_only=True):
                values = ["" if cell is None else str(cell) for cell in row]
                if any(values):
                    lines.append(", ".join(values))
        preview = "\n".join(lines)[:MAX_PREVIEW_CHARS]
        return preview, f"XLSX with {len(workbook.sheetnames)} sheets.", "partial", None
    except Exception as exc:
        return "", f"XLSX extraction unavailable: {exc}", "metadata_only", None
